import openpyxl

def read_and_edit_excel_file():
    # Open xlsx file
    inv_file = openpyxl.load_workbook("inventory.xlsx")

    prod_list = inv_file.active

    # Initialize dictionaries
    comp_prod_dict = {}
    prod_less_than_ten = {}
    total_inv_value = {}

    # Loop through rows in xlsx file
    for prod_row in range(2, prod_list.max_row + 1):
        supplier_name = prod_list.cell(prod_row, 4).value
        inventory = prod_list.cell(prod_row, 2).value
        price = prod_list.cell(prod_row, 3).value
        prod_id = prod_list.cell(prod_row, 1).value
        inv_value = inventory * price
        inventory_price = prod_list.cell(prod_row, 5)

        if supplier_name in comp_prod_dict.keys():
            comp_prod_dict[supplier_name] = comp_prod_dict[supplier_name] + 1
        else:
            comp_prod_dict[supplier_name] = 1
        if inventory < 10:
            prod_less_than_ten[prod_id] = inventory
        if supplier_name in total_inv_value:
            total_inv_value[supplier_name] = total_inv_value[supplier_name] + inv_value
        else:
            total_inv_value[supplier_name] = inv_value

        # Create a new column for each inventory total price
        inventory_price.value = inv_value

    print(f"company products: {comp_prod_dict}")
    print(f"products less than 10: {prod_less_than_ten}")
    print(f"total inventory value: {total_inv_value}")

    # Save newly created file with the new column
    inv_file.save("new_inventory.xlsx")

if __name__ == "__main__":

    read_and_edit_excel_file()